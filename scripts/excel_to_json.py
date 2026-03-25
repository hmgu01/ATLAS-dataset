"""
excel_to_json.py
────────────────────────────────────────────────────────────────
ATLAS Dataset — Excel → JSON Converter
Converts architectural spatial relationship Excel files to the
ATLAS dataset JSON schema (horizontal + vertical graphs + areas).

Usage:
    # Single file
    python excel_to_json.py --input data/1_차이커뮤니케이션_사옥.xlsx --output data/json/

    # Batch (all xlsx in a folder)
    python excel_to_json.py --batch --input data/xlsx/ --output data/json/

Output per building:
    {building_id}/
        horizontal_graph.json   ← intra-floor adjacency per floor
        vertical_graph.json     ← inter-floor core continuity
        area_ratios.json        ← space areas and floor ratios
        metadata.json           ← building summary

Column schema (both sheets):
    A(0)  : 건축물+층수  (building + floor ID)
    B(1)  : 공간이름     (space name / class label)
    C(2)  : 공간CLASS    (node class ID)
    D(3)  : N            (instance index)
    E(4)  : 인접공간이름  (adjacent space name)
    F(5)  : 인접CLASS    (adjacent class ID)
    G(6)  : 인접N        (adjacent instance index)
    J(9)  : 인접성       (edge weight, 0.0–1.0)

Area info (horizontal sheet only, right side):
    R(17) : 건축물+층수  (building + floor)
    S(18) : 공간+넘버    (space instance name)
    T(19) : 면적(㎡)     (area in sqm)
────────────────────────────────────────────────────────────────
"""

import argparse
import csv
import json
import os
import re
from collections import defaultdict
from pathlib import Path

import openpyxl


# ── Node class name → ID mapping (22 classes) ──────────────────
NODE_CLASS_MAP = {
    "임대": 1, "메인도로": 2, "복도": 3, "방풍실": 4,
    "물탱크실": 5, "홀": 6, "설비": 7, "E.V": 8,
    "계단실": 9, "화장실": 10, "관리실": 11, "창고": 12,
    "테라스": 13, "부도로": 14, "지하주차장 진입로": 15,
    "공지": 16, "녹지": 17, "주변매스": 18, "주차장": 19,
    "기계실": 20, "방재실": 21, "전기실": 22,
    "PIT": 23, "썬큰": 24, "DA": 25,
}

# ── Korean → English class name mapping ────────────────────────
NODE_CLASS_EN = {
    "임대":           "Leasable Unit",
    "메인도로":        "Main Road",
    "복도":           "Corridor",
    "방풍실":         "Vestibule",
    "물탱크실":        "Water Tank Room",
    "홀":            "Hall / Lobby",
    "설비":           "MEP / Utility Room",
    "E.V":           "Elevator",
    "계단실":         "Staircase",
    "화장실":         "Restroom",
    "관리실":         "Management Room",
    "창고":           "Storage",
    "테라스":         "Terrace",
    "부도로":         "Secondary Road",
    "지하주차장 진입로": "Underground Parking Entrance",
    "공지":           "Open Space",
    "녹지":           "Green Space",
    "주변매스":        "Surrounding Mass",
    "주차장":         "Parking Lot",
    "기계실":         "Mechanical Room",
    "방재실":         "Disaster Prevention Room",
    "전기실":         "Electrical Room",
    "PIT":            "Pit / Sunken Area",
    "썬큰":           "Sunken Court",
    "DA":             "Dry Area",
}

# ── External context node classes ──────────────────────────────
EXTERNAL_CONTEXT_CLASSES = {2, 14, 16, 17, 18}

# ── Vertical core classes (mandatory continuity) ───────────────
VERTICAL_CORE_CLASSES = {8, 9}   # E.V, 계단실

# ── Floor sort order ────────────────────────────────────────────
FLOOR_ORDER = {
    "B3": -3, "B2": -2, "B1": -1,
    "1": 1, "2": 2, "3": 3, "4": 4,
    "5": 5, "6": 6, "7": 7, "8": 8,
    "9": 9, "10": 10, "11": 11, "12": 12,
}


def floor_sort_key(floor_str):
    """Sort floor labels: B2 < B1 < 1F < 2F ..."""
    m = re.search(r"[+](.+)$", floor_str)
    tag = m.group(1) if m else floor_str
    return FLOOR_ORDER.get(tag, 99)


def parse_building_floor(cell_value):
    """
    '차이커뮤니케이션 사옥+B2' → ('차이커뮤니케이션 사옥', 'B2')
    """
    if not cell_value:
        return None, None
    parts = str(cell_value).rsplit("+", 1)
    if len(parts) == 2:
        return parts[0].strip(), parts[1].strip()
    return str(cell_value).strip(), "UNK"


def safe_int(val, default=None):
    """공백/None/비정수 값을 안전하게 int 변환. 실패 시 default 반환."""
    if val is None:
        return default
    s = str(val).strip()
    if not s:
        return default
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return default


def make_node_id(class_name, instance_n):
    """Stable node ID: 'E.V_1', '계단실_3', etc."""
    return f"{class_name}_{instance_n}"


def load_area_info(ws_horizontal):
    """
    Read area table from right side of horizontal sheet.
    Returns: { 'building+floor': { 'space_name_instance': area_sqm } }
    """
    areas = defaultdict(dict)
    rows = list(ws_horizontal.iter_rows(values_only=True))
    # Area columns: R(17), S(18), T(19) — 0-indexed
    for row in rows[1:]:  # skip header row 0
        bldg_floor = row[17]
        space_num  = row[18]
        area       = row[19]
        if bldg_floor and space_num and area and isinstance(area, (int, float)):
            areas[str(bldg_floor)][str(space_num)] = round(float(area), 3)
    return areas


def build_horizontal_graph(ws_horizontal):
    """
    Parse horizontal (intra-floor) adjacency sheet.
    Returns: { floor_key: { nodes: [...], edges: [...] } }
    """
    floors = defaultdict(lambda: {"nodes": {}, "edges": []})
    rows = list(ws_horizontal.iter_rows(values_only=True))

    # Track seen edges to avoid duplicates (undirected graph)
    seen_edges = defaultdict(set)

    for row in rows[3:]:  # rows 0-2 are headers
        bldg_floor = row[0]
        if not bldg_floor:
            continue

        src_name    = row[1]
        src_class   = row[2]
        src_n       = row[3]
        tgt_name    = row[4]
        tgt_class   = row[5]
        tgt_n       = row[6]

        # 필수값 빈칸 체크 → 행 스킵
        if not str(src_name).strip() or safe_int(src_class) is None or safe_int(src_n) is None:
            continue
        if not str(tgt_name).strip() or safe_int(tgt_class) is None or safe_int(tgt_n) is None:
            continue
        weight      = row[9]

        if not all([src_name, src_class, src_n, tgt_name, tgt_class, tgt_n]):
            continue
        if weight is None:
            continue

        floor_key = str(bldg_floor)
        floor_data = floors[floor_key]

        # Register source node
        src_id = make_node_id(src_name, src_n)
        if src_id not in floor_data["nodes"]:
            floor_data["nodes"][src_id] = {
                "id": src_id,
                "class_id": safe_int(src_class),
                "class_name": str(src_name),
                "class_name_en": NODE_CLASS_EN.get(str(src_name), str(src_name)),
                "instance_index": safe_int(src_n),
                "is_external_context": safe_int(src_class) in EXTERNAL_CONTEXT_CLASSES,
                "is_vertical_core": safe_int(src_class) in VERTICAL_CORE_CLASSES,
            }

        # Register target node
        tgt_id = make_node_id(tgt_name, tgt_n)
        if tgt_id not in floor_data["nodes"]:
            floor_data["nodes"][tgt_id] = {
                "id": tgt_id,
                "class_id": safe_int(tgt_class),
                "class_name": str(tgt_name),
                "class_name_en": NODE_CLASS_EN.get(str(tgt_name), str(tgt_name)),
                "instance_index": safe_int(tgt_n),
                "is_external_context": safe_int(tgt_class) in EXTERNAL_CONTEXT_CLASSES,
                "is_vertical_core": safe_int(tgt_class) in VERTICAL_CORE_CLASSES,
            }

        # Register edge (undirected: store canonical order)
        edge_key = tuple(sorted([src_id, tgt_id]))
        if edge_key not in seen_edges[floor_key]:
            seen_edges[floor_key].add(edge_key)
            floor_data["edges"].append({
                "source": src_id,
                "target": tgt_id,
                "weight": round(float(weight), 2),
            })

    return floors


def build_vertical_graph(ws_vertical):
    """
    Parse vertical (inter-floor) adjacency sheet.
    Returns: [ { source_floor, target_floor, source_node, target_node, weight } ]
    """
    edges = []
    seen = set()
    rows = list(ws_vertical.iter_rows(values_only=True))

    for row in rows[3:]:
        bldg_floor = row[0]
        if not bldg_floor:
            continue

        src_name  = row[1]
        src_class = row[2]
        src_n     = row[3]
        tgt_name  = row[4]
        tgt_class = row[5]
        tgt_n     = row[6]

        # 필수값 빈칸 체크
        if not str(src_name).strip() or safe_int(src_class) is None or safe_int(src_n) is None:
            continue
        if not str(tgt_name).strip() or safe_int(tgt_class) is None or safe_int(tgt_n) is None:
            continue
        weight    = row[9]

        if not all([src_name, src_class, src_n, tgt_name, tgt_class, tgt_n]):
            continue
        if weight is None:
            continue

        bldg, src_floor = parse_building_floor(bldg_floor)
        src_id = make_node_id(src_name, src_n)
        tgt_id = make_node_id(tgt_name, tgt_n)

        edge_key = (str(bldg_floor), src_id, tgt_id)
        if edge_key not in seen:
            seen.add(edge_key)
            edges.append({
                "source_floor": src_floor,
                "source_node":  src_id,
                "source_class_id": safe_int(src_class),
                "target_node":  tgt_id,
                "target_class_id": safe_int(tgt_class),
                "weight": round(float(weight), 2),
                "is_mandatory_continuity": round(float(weight), 2) == 1.0,
            })

    return edges


def compute_area_ratios(area_info):
    """
    Compute area ratios per floor (space_area / total_floor_area).
    Returns: { floor_key: { space_name: { area_sqm, area_ratio } } }
    """
    result = {}
    for floor_key, spaces in area_info.items():
        total = sum(v for v in spaces.values() if isinstance(v, (int, float)))
        result[floor_key] = {}
        for space_name, area in spaces.items():
            result[floor_key][space_name] = {
                "area_sqm": area,
                "area_ratio": round(area / total, 4) if total > 0 else 0.0,
                "total_floor_area_sqm": round(total, 3),
            }
    return result



# ── External context face assignment ──────────────────────────
SLAB_W_FACE   = {2:28, 14:20, 16:25, 17:25, 18:58}
LAYER_ORD_FACE= {16:0, 17:0, 2:1, 14:1, 18:2}
MASS_FACE_MAP = {1:"S", 2:"W", 3:"N", 4:"E"}
ROAD_FACE_MAP = {"메인도로":"E","메인 도로":"E","부도로":"N",
                 "주차장 진입로":"E","지하주차장 진입로":"E"}

def assign_faces(floor_nodes: dict, floor_edges: list,
                 int_positions: dict) -> dict:
    """
    3-step face assignment for external context nodes in one floor.
    Returns: { node_id: face_str }  where face_str ∈ {"N","S","E","W"}

    int_positions: { node_id: (x, z) } spring-layout coordinates.
    If not available (e.g. no networkx), Step 3 falls back to name pattern.
    """
    ext_nodes = {nid: n for nid, n in floor_nodes.items()
                 if n["is_external_context"]}
    face_result = {}

    # Step 1: name pattern (definitive)
    for nid, n in ext_nodes.items():
        name = n["class_name"]; num = n["instance_index"]
        for road_kw, f in ROAD_FACE_MAP.items():
            if road_kw in name:
                face_result[nid] = f; break
        if nid not in face_result and "주변매스" in name:
            face_result[nid] = MASS_FACE_MAP.get(num, "N")

    # Step 2: propagate via ext→ext edges (3 rounds)
    for _ in range(3):
        for e in floor_edges:
            for src, tgt in [(e["source"], e["target"]),
                             (e["target"], e["source"])]:
                if (src in face_result and tgt in ext_nodes
                        and tgt not in face_result):
                    face_result[tgt] = face_result[src]

    # Step 3: connected internal node position (if available)
    def face_of(x, z):
        if abs(z) > abs(x): return "N" if z > 0 else "S"
        return "E" if x > 0 else "W"

    for nid, n in ext_nodes.items():
        if nid in face_result:
            continue
        conn_pts = []
        for e in floor_edges:
            other = None
            if e["source"] == nid:   other = e["target"]
            elif e["target"] == nid: other = e["source"]
            if other and other in int_positions and other not in ext_nodes:
                conn_pts.append(int_positions[other])
        if conn_pts:
            ax = sum(p[0] for p in conn_pts) / len(conn_pts)
            az = sum(p[1] for p in conn_pts) / len(conn_pts)
            face_result[nid] = face_of(-ax, -az)
        else:
            # Fallback: number pattern
            name = n["class_name"]; num = n["instance_index"]
            if "공지" in name:
                face_result[nid] = {1:"S",2:"W",3:"N",4:"E"}.get(num,"N")
            else:
                face_result[nid] = "N"

    return face_result


def compute_ext_offsets(ext_nodes: dict, face_map: dict) -> dict:
    """Compute radial offset per external node for 3D placement."""
    from collections import defaultdict
    face_groups = defaultdict(list)
    for nid, n in ext_nodes.items():
        face = face_map.get(nid, "N")
        face_groups[face].append(
            (LAYER_ORD_FACE.get(n["class_id"], 0),
             n["class_id"], n["instance_index"], nid))
    offsets = {}
    for face, items in face_groups.items():
        items.sort(); off = 72
        for _, cls, _, nid in items:
            w = SLAB_W_FACE.get(cls, 25)
            off += w / 2; offsets[nid] = round(off, 1); off += w / 2
    return offsets


def convert_excel_to_json(xlsx_path: Path, output_dir: Path, anonymize_id: str = None):
    """
    Main conversion function.
    Reads one Excel file and writes 4 JSON files to output_dir/{building_id}/
    """
    print(f"  Processing: {xlsx_path.name}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    if "수평" not in wb.sheetnames or "수직" not in wb.sheetnames:
        print(f"  ⚠️  Skipping {xlsx_path.name}: '수평' or '수직' sheet not found")
        return

    ws_h = wb["수평"]
    ws_v = wb["수직"]

    # Parse all data
    h_floors   = build_horizontal_graph(ws_h)
    v_edges    = build_vertical_graph(ws_v)
    area_raw   = load_area_info(ws_h)
    area_ratios = compute_area_ratios(area_raw)

    # Derive building name from filename or first data cell
    building_name = xlsx_path.stem  # e.g. '1_차이커뮤니케이션_사옥'
    if anonymize_id:
        building_id = anonymize_id          # e.g. 'building_004'
    else:
        building_id = re.sub(r"[^\w가-힣]", "_", building_name)

    # Sort floors
    sorted_floors = sorted(h_floors.keys(), key=floor_sort_key)

    def anon_fk(floor_key: str) -> str:
        """Replace 'BuildingName+FL' → 'building_NNN+FL' when anonymizing."""
        if not anonymize_id:
            return floor_key
        _, fl = parse_building_floor(floor_key)
        return f"{anonymize_id}+{fl}"

    # ── Try spring layout for face Step-3 positions ─────────────
    int_positions_by_floor = {}
    try:
        import networkx as nx
        for fk in sorted_floors:
            fd = h_floors[fk]
            int_ids = [n["id"] for n in fd["nodes"].values()
                       if not n["is_external_context"]]
            if not int_ids: continue
            G = nx.Graph(); G.add_nodes_from(int_ids)
            for e in fd["edges"]:
                s, t = e["source"], e["target"]
                if s in int_ids and t in int_ids:
                    G.add_edge(s, t, weight=max(float(e["weight"]), 0.1))
            try:
                pos = nx.spring_layout(G, k=0.35, iterations=80, seed=42)
                int_positions_by_floor[fk] = {
                    nid: (round(px*72, 1), round(pz*72, 1))
                    for nid, (px, pz) in pos.items()}
            except Exception:
                pass
    except ImportError:
        pass  # networkx optional — face Step 3 falls back to name pattern

    # ── Build horizontal_graph.json ──────────────────────────────
    horizontal_output = []
    for floor_key in sorted_floors:
        _, floor_label = parse_building_floor(floor_key)
        floor_data = h_floors[floor_key]
        int_pos = int_positions_by_floor.get(floor_key, {})

        # Assign face & offset to external context nodes
        face_map = assign_faces(
            floor_data["nodes"], floor_data["edges"], int_pos)
        ext_off = compute_ext_offsets(
            {nid: n for nid, n in floor_data["nodes"].items()
             if n["is_external_context"]}, face_map)

        nodes_out = []
        for n in floor_data["nodes"].values():
            nd = dict(n)
            if n["is_external_context"]:
                nd["face"]               = face_map.get(n["id"], "N")
                nd["offset"]             = ext_off.get(n["id"], 97.0)
                nd["is_surrounding_mass"]= n["class_id"] == 18
            nodes_out.append(nd)

        horizontal_output.append({
            "floor_key":  anon_fk(floor_key),
            "floor_label": floor_label,
            "num_nodes":  len(floor_data["nodes"]),
            "num_edges":  len(floor_data["edges"]),
            "nodes":      nodes_out,
            "edges":      floor_data["edges"],
        })

    # ── Build vertical_graph.json ────────────────────────────────
    vertical_output = {
        "building_id": building_id,
        "total_inter_floor_edges": len(v_edges),
        "mandatory_continuity_edges": sum(1 for e in v_edges if e["is_mandatory_continuity"]),
        "edges": v_edges,
    }

    # ── Build area_ratios.json ───────────────────────────────────
    area_output = {
        "building_id": building_id,
        "floors": {
            anon_fk(floor_key): {
                "floor_label": parse_building_floor(floor_key)[1],
                "spaces": area_ratios.get(floor_key, {}),
            }
            for floor_key in sorted_floors
        }
    }

    # ── Build metadata.json ──────────────────────────────────────
    all_node_classes = set()
    total_nodes = 0
    total_edges = 0
    for fd in h_floors.values():
        for node in fd["nodes"].values():
            all_node_classes.add(node["class_id"])
        total_nodes += len(fd["nodes"])
        total_edges += len(fd["edges"])

    # Warn about undefined class IDs (outside 1~22)
    unknown_classes = sorted(all_node_classes - set(range(1, 26)))  # 1~25 valid
    if unknown_classes:
        print(f"  ⚠️  Unknown class IDs found: {unknown_classes} "
              f"— check Excel CLASS column (valid: 1~22)")

    metadata = {
        "building_id": building_id,
        "source_file": xlsx_path.name if not anonymize_id else "[ANONYMIZED]",
        "num_floors": len(sorted_floors),
        "floor_labels": [parse_building_floor(f)[1] for f in sorted_floors],
        "total_nodes": total_nodes,
        "total_horizontal_edges": total_edges,
        "total_vertical_edges": len(v_edges),
        "node_classes_present": sorted(list(all_node_classes)),
        "has_basement": any("B" in f for f in sorted_floors),
        "schema_version": "1.0",
    }

    # ── Write output files ───────────────────────────────────────
    out_dir = output_dir / building_id
    out_dir.mkdir(parents=True, exist_ok=True)

    files = {
        "horizontal_graph.json": horizontal_output,
        "vertical_graph.json":   vertical_output,
        "area_ratios.json":      area_output,
        "metadata.json":         metadata,
    }

    for filename, data in files.items():
        out_path = out_dir / filename
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"  ✅ {building_id}/ → {len(sorted_floors)} floors, "
          f"{total_nodes} nodes, {total_edges} h-edges, {len(v_edges)} v-edges")
    return metadata


def extract_number(stem: str) -> int:
    """Extract leading number from filename: '4_더_스타일_빌딩' → 4."""
    m = re.match(r"^(\d+)", stem)
    return int(m.group(1)) if m else 0


def make_anonymized_id(number: int) -> str:
    """'4' → 'building_004'."""
    return f"building_{number:03d}"


def main():
    parser = argparse.ArgumentParser(
        description="ATLAS Dataset: Excel → JSON converter"
    )
    parser.add_argument("--input",  required=True,
                        help="Input .xlsx file or folder (use with --batch)")
    parser.add_argument("--output", required=True,
                        help="Output directory for JSON files")
    parser.add_argument("--batch",  action="store_true",
                        help="Process all .xlsx files in --input folder")
    parser.add_argument("--anonymize", action="store_true",
                        help="Replace building names with building_NNN IDs. "
                             "Saves a mapping.csv to --output (keep private).")
    args = parser.parse_args()

    output_dir = Path(args.output)
    output_dir.mkdir(parents=True, exist_ok=True)

    if args.batch:
        input_dir = Path(args.input)
        xlsx_files = sorted(input_dir.glob("*.xlsx"))
        print(f"\nBatch mode: {len(xlsx_files)} files found in {input_dir}")
        if args.anonymize:
            print("Anonymization: ON  (mapping.csv will be written)\n")
        else:
            print()

        summaries = []
        mapping_rows = []   # (anon_id, original_filename)

        for f in xlsx_files:
            anon_id = None
            if args.anonymize:
                num    = extract_number(f.stem)
                anon_id = make_anonymized_id(num)
                mapping_rows.append((anon_id, f.name, f.stem))
                print(f"  {f.name}  →  {anon_id}/")

            meta = convert_excel_to_json(f, output_dir, anonymize_id=anon_id)
            if meta:
                summaries.append(meta)

        # Write batch summary
        summary_path = output_dir / "dataset_summary.json"
        with open(summary_path, "w", encoding="utf-8") as fp:
            json.dump({
                "total_buildings": len(summaries),
                "anonymized": args.anonymize,
                "buildings": summaries
            }, fp, ensure_ascii=False, indent=2)
        print(f"\n📊 Summary written → {summary_path}")

        # Write mapping.csv (only when anonymize)
        if args.anonymize and mapping_rows:
            mapping_path = output_dir / "mapping.csv"
            with open(mapping_path, "w", newline="", encoding="utf-8") as fp:
                writer = csv.writer(fp)
                writer.writerow(["building_id", "original_filename", "original_stem"])
                writer.writerows(mapping_rows)
            print(f"🔒 Mapping written  → {mapping_path}  ← keep this PRIVATE")

    else:
        # Single file
        anon_id = None
        if args.anonymize:
            num    = extract_number(Path(args.input).stem)
            anon_id = make_anonymized_id(num)
            print(f"Anonymization: {Path(args.input).name}  →  {anon_id}/")
            # Write single-file mapping
            mapping_path = output_dir / "mapping.csv"
            with open(mapping_path, "w", newline="", encoding="utf-8") as fp:
                writer = csv.writer(fp)
                writer.writerow(["building_id", "original_filename", "original_stem"])
                writer.writerow([anon_id, Path(args.input).name, Path(args.input).stem])
            print(f"🔒 Mapping written  → {mapping_path}  ← keep this PRIVATE")

        convert_excel_to_json(Path(args.input), output_dir, anonymize_id=anon_id)

    print("\nDone.")


if __name__ == "__main__":
    main()
